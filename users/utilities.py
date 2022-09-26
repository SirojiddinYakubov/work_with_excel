from typing import Union
import openpyxl
import os
from io import BytesIO
from django.db.models import QuerySet
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import BLACK
from conf import settings
from datetime import datetime
from users.models import CustomUser
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
import smtplib


def get_file_path(file_name: str) -> str:
    return os.path.join(settings.BASE_DIR, file_name)


def filter_user_data(data: dict) -> bool:
    if len(str(data['person_id'])) != 14:
        return False
    # if len(data['first_name']) < 5:
    #     return False
    return all([bool(value) for value in data.values()])


# str_to_date = lambda value: datetime.strptime(value, "%d.%m.%Y").date()
# get_gender_value = lambda value: CustomUser.GENDER_MALE if value == 'erkak' else CustomUser.GENDER_FEMALE


def get_real_value(data: dict) -> dict:
    if data['birthday']:
        data['birthday'] = datetime.strptime(data['birthday'], "%d.%m.%Y").date()
    if data['gender']:
        data['gender'] = CustomUser.GENDER_MALE if data['gender'] == 'erkak' else CustomUser.GENDER_FEMALE
    return data


def users_data_parse_from_excel(file_name: str) -> Union[list[dict], tuple[dict]]:
    file_path = get_file_path(file_name=file_name)
    book = openpyxl.open(filename=file_path, read_only=True)
    sheet = book.active

    datas = []
    keys = ['person_id', 'first_name', 'last_name', 'email', 'gender', 'salary', 'birthday']
    for row in sheet.iter_rows(min_col=2, max_col=8, min_row=2, max_row=sheet.max_row + 1):
        data = dict(zip(keys, (cell.value for cell in row)))
        datas.append(data)

    return list(map(get_real_value, filter(filter_user_data, datas)))


def write_users_data_to_excel(file_name: str, sheet_name: str, users_data: Union[tuple, list, QuerySet]):
    # file_path = get_file_path(file_name=file_name)
    # book = openpyxl.open(filename=file_path)
    #
    # if not sheet_name in book.sheetnames:
    #     book.create_sheet(sheet_name)
    # sheet = book[sheet_name]

    book = openpyxl.Workbook()
    sheet = book.active

    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15

    # set headers
    headers = ['ID', 'person_id', 'email', 'last_name', 'first_name', 'salary', 'birthday']
    for i_col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=i_col, value=header)
        cell.font = Font(size=11, color=BLACK, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i_row, user_data in enumerate(users_data, start=2):
        for i_col, value in enumerate(user_data, start=1):
            cell = sheet.cell(row=i_row, column=i_col, value=value)

            if i_col == 6:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    book.save(file_name)


def users_salary_report_build(users: QuerySet) -> BytesIO:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'отчёт'

    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    headers = ['№', 'Электронная почта', 'Фамилия', 'Имя', 'Зарплата']
    for i_col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=i_col, value=header)
        cell.font = Font(size=11, color=BLACK, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i_row, user_data in enumerate(users, start=2):
        user_data = [i_row - 1] + list(user_data)
        for i_col, value in enumerate(user_data, start=1):
            cell = sheet.cell(row=i_row, column=i_col, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if i_row == len(users) + 1 and i_col == len(user_data):
                char = get_column_letter(i_col)

                cell = sheet.cell(row=i_row + 1, column=i_col, value=f"=SUM({char}2:{char + str(i_row)})")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=14, color=BLACK, bold=True)

                next_row = str(len(users) + 2)
                sheet['A' + next_row] = 'Итого'
                sheet.merge_cells(f"A{next_row}:D{next_row}")

    return BytesIO(save_virtual_workbook(workbook))


def users_salary_report_file_build_for_email_send(users: QuerySet) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'отчёт'

    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    headers = ['№', 'Электронная почта', 'Фамилия', 'Имя', 'Зарплата']
    for i_col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=i_col, value=header)
        cell.font = Font(size=11, color=BLACK, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i_row, user_data in enumerate(users, start=2):
        user_data = [i_row - 1] + list(user_data)
        for i_col, value in enumerate(user_data, start=1):
            cell = sheet.cell(row=i_row, column=i_col, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if i_row == len(users) + 1 and i_col == len(user_data):
                char = get_column_letter(i_col)

                cell = sheet.cell(row=i_row + 1, column=i_col, value=f"=SUM({char}2:{char + str(i_row)})")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=14, color=BLACK, bold=True)

                next_row = str(len(users) + 2)
                sheet['A' + next_row] = 'Итого'
                sheet.merge_cells(f"A{next_row}:D{next_row}")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class SmtpServer:

    def __init__(self, recipients: list, file: bytes):
        self.recipients = recipients
        self.file = file
        self.sender = settings.EMAIL_HOST_USER
        self.sender_password = settings.EMAIL_HOST_PASSWORD
        self.host = settings.EMAIL_HOST
        self.port = settings.EMAIL_PORT
        self.filename = 'report.xlsx'
        self.subject = 'Salary Report'

    def send(self):
        msg = MIMEMultipart()
        server = smtplib.SMTP(self.host, self.port)
        server.starttls()
        server.login(user=self.sender, password=self.sender_password)

        msg['Subject'] = self.subject
        msg["From"] = self.sender
        msg["To"] = ", ".join(self.recipients)

        xlsx = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        xlsx.set_payload(self.file)
        encoders.encode_base64(xlsx)
        xlsx.add_header('Content-Disposition', 'attachment', filename=self.filename)
        msg.attach(xlsx)

        server.sendmail(self.sender, self.recipients, msg.as_string())
        server.quit()
